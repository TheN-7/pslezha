import sqlite3
from datetime import date, datetime
from pathlib import Path
from re import sub

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "Dogana_Template.xlsm"
DB_PATH = BASE_DIR / "dogana.db"
SHEET_NAME = "Master_Database"
BATCH_SIZE = 1000


def slugify(value):
    value = str(value or "").strip().lower()
    value = value.replace("'", "")
    value = sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_") or "column"


def unique_names(headers):
    seen = {}
    names = []
    for header in headers:
        name = slugify(header)
        count = seen.get(name, 0)
        seen[name] = count + 1
        names.append(name if count == 0 else f"{name}_{count + 1}")
    return names


def clean_value(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    return value


def create_schema(db, columns, headers):
    db.execute("DROP TABLE IF EXISTS voters")
    db.execute("DROP TABLE IF EXISTS column_metadata")

    column_defs = ", ".join(f'"{column}" TEXT' for column in columns)
    db.execute(
        f"""
        CREATE TABLE voters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {column_defs}
        )
        """
    )

    db.execute(
        """
        CREATE TABLE column_metadata (
            column_name TEXT PRIMARY KEY,
            original_header TEXT NOT NULL,
            ordinal INTEGER NOT NULL
        )
        """
    )
    db.executemany(
        """
        INSERT INTO column_metadata (column_name, original_header, ordinal)
        VALUES (?, ?, ?)
        """,
        [(column, str(header), index) for index, (column, header) in enumerate(zip(columns, headers), 1)],
    )

    db.execute(
        """
        CREATE VIEW voter_search AS
        SELECT
            id,
            uniqueid,
            first_name,
            last_name,
            fathers_name,
            mothers_name,
            birthday,
            age,
            gender,
            admin_unit,
            shtab,
            os,
            qv,
            qv_ranking_no,
            building_no,
            political_preference,
            phone_number,
            emigrant
        FROM voters
        """
    )


def create_indexes(db):
    indexes = [
        ("idx_voters_uniqueid", "uniqueid"),
        ("idx_voters_first_name", "first_name"),
        ("idx_voters_last_name", "last_name"),
        ("idx_voters_fathers_name", "fathers_name"),
        ("idx_voters_admin_unit", "admin_unit"),
        ("idx_voters_qv", "qv"),
        ("idx_voters_shtab", "shtab"),
    ]

    for index_name, column in indexes:
        db.execute(f'CREATE INDEX IF NOT EXISTS "{index_name}" ON voters ("{column}")')


def convert():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Missing workbook: {EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True, keep_vba=False)
    worksheet = workbook[SHEET_NAME]

    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(rows)]
    columns = unique_names(headers)

    with sqlite3.connect(DB_PATH) as db:
        create_schema(db, columns, headers)

        placeholders = ", ".join("?" for _ in columns)
        column_list = ", ".join(f'"{column}"' for column in columns)
        insert_sql = f"INSERT INTO voters ({column_list}) VALUES ({placeholders})"

        batch = []
        imported = 0
        for row in rows:
            if not any(value not in (None, "") for value in row):
                continue

            clean_row = [clean_value(value) for value in row[: len(columns)]]
            if len(clean_row) < len(columns):
                clean_row.extend([None] * (len(columns) - len(clean_row)))

            batch.append(clean_row)
            if len(batch) >= BATCH_SIZE:
                db.executemany(insert_sql, batch)
                imported += len(batch)
                batch.clear()

        if batch:
            db.executemany(insert_sql, batch)
            imported += len(batch)

        create_indexes(db)

    print(f"Created {DB_PATH.name}")
    print(f"Imported rows: {imported}")
    print(f"Columns: {len(columns)}")


if __name__ == "__main__":
    convert()
